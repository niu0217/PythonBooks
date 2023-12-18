import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font

# wb = openpyxl.load_workbook('./sources/example.xlsx')
# print(wb.sheetnames)  # 打印表名，是个列表
# sheet = wb['Sheet1']  # Get a sheet from the workbook.

# 打印第1行第A列的第一个数据
# content1 = sheet['A1']
# print(f'Row: {content1.row} Column: {content1.column}  Value: {content1.value}')

# 打印第2列的奇数行的数据
# for i in range(1, 8, 2):
#     print(i, sheet.cell(row=i, column=2).value)

# 打印Sheet1的最大行和列
# print(f"Sheet1's max row: {sheet.max_row}")
# print(f"Sheet1's max column: {sheet.max_column}")

# 列字母和数字之间的转换
# print(get_column_letter(27))  # 列数字转换成字母
# print(column_index_from_string('AD'))  # 列字母转换成数字

# 打印3行3列的数据
# print(tuple(sheet['A1':'C3']))
# for rowOfCellObjects in sheet['A1':'C3']:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, cellObj.value)
#     print('')

# 打印第2列的所有数据
# print(list(sheet.columns)[1])
# for cellObj in list(sheet.columns)[1]:
#     print(cellObj.value)

# 创建并保存excel文档
# wb = openpyxl.load_workbook('sources/example.xlsx')
# sheet = wb.active
# sheet.title = 'Spam Spam Spam'
# wb.save('example.xlsx')

# 创建和删除工作表
# wb = openpyxl.Workbook()
# print(wb.sheetnames)
# wb.create_sheet()
# print(wb.sheetnames)
# wb.create_sheet(index=0, title='First Sheet')
# print(wb.sheetnames)
# wb.create_sheet(index=2, title='Middle Sheet')
# print(wb.sheetnames)
# del wb['Middle Sheet']
# del wb['Sheet1']
# print(wb.sheetnames)
# wb.save('nihao.xlsx')

# 将值写入单元格
# wb = openpyxl.Workbook()
# print(wb.sheetnames)
# sheet = wb['Sheet']
# sheet['A1'] = 'Hello niu0217'
# print(sheet['A1'].value)
# # wb.save('nihao.xlsx')

# 设置单元格的字体
# wb = openpyxl.Workbook()
# sheet = wb.active
# italic24Font = Font(size=24, italic=True)
# sheet['A1'].font = italic24Font
# sheet['A1'] = 'Hello niu'
# sheet['B9'] = '=SUM(B1:B8)'  # 设置公式
# wb.save('style.xlsx')

# 调整行和列
# wb = openpyxl.Workbook()
# sheet = wb.active
# 调整高度和宽度
# sheet['A1'] = 'Tall row'
# sheet['B2'] = 'Ni hao'
# sheet.row_dimensions[1].height = 70
# sheet.column_dimensions['B'].width = 20
# wb.save('dimensions.xlsx')
# 合并拆分单元格
# sheet.merged_cells('A1:D3')
# sheet['A1'] = 'Tall row'
# sheet.unmerge_cells('A1:D3')
# wb.save('merged.xlsx')
# 冻结窗格
# sheet.freeze_panes = 'A2'  # Freeze the rows above A2
# wb.save('freeze.xlsx')


