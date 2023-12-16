#! python3
# readCensusExcel.py
# Tabulates population and number of census tracts for
# each contry.

import openpyxl
import pprint

print('Opening workbook...')
wb = openpyxl.load_workbook('sources/censuspopdata.xlsx')
# print(wb.sheetnames)
sheet = wb['Population by Census Tract']
countryData = {}

# Fill in countryData with each country's population and tracts
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one census tract.
    state = sheet['B' + str(row)].value  # 州的简称
    country = sheet['C' + str(row)].value  # 县的名称
    pop = sheet['D' + str(row)].value  # 普查区的人口

    # Make sure the key for this state exists
    countryData.setdefault(state, {})
    # Make sure the key for this country in this state exists
    countryData[state].setdefault(country, {'tracts': 0, 'pop': 0})
    countryData[state][country]['tracts'] += 1
    countryData[state][country]['pop'] += int(pop)

# Open a new text file and write the contents of countryData to it.
print('Writing results...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countryData))
resultFile.close()
print('Done')
