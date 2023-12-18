#! python3
# updateProduce.py
# 修正A列中的值

import openpyxl

wb = openpyxl.load_workbook('sources/produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

# Loop through the rows and updates the prices
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updateProduceSales.xlsx')
